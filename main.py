from persoenliches import *
from progressing import ProgressIndicator
from openpyxl import load_workbook
import mailing



def run():

    # today = date.today()

    # falls vorlage nicht existiert, wird standard benutzt
    try:
        wb = load_workbook(WORKBOOK_PATH)
    except:
        wb = load_workbook('/vorlage/Stundenzettel-VORLAGE.xlsx')

    if not os.path.exists(OUTPUT_PATH):
        os.makedirs(OUTPUT_PATH)

    # Arbeitsblatt auswählen
    worksheet = wb.active

    # Abfrage, ob heute schon Freitag ist
    # is_friday = TODAY == FIRST_DAY + timedelta(days=4)

    # if is_friday:
    #     print(f'Heute ist Freitag, den {today}')
    # else:
    for NAMENSFELD in DATEN:
        # nimmt die Felder aus der Excell und upgraded sie mit die der in der DATEN Dict.
        worksheet[NAMENSFELD] = DATEN[NAMENSFELD]
    wb.save(WORKBOOK_PATH)

    # Erstelle einen Dateinamen für das PDF
    pdf_file_name = f'KW{WEEKNUMBER}_{NACHNAME.replace(" ", "_")}_{YEAR}.pdf'
    # Erstelle den Ausgabepfad für das PDF
    output_path = f'{OUTPUT_PATH}{pdf_file_name}'


    save_pdf(WORKBOOK_PATH, output_path)
    # mailing.send_mail(SEND_MAIL_FROM, SEND_MAIL_TO, f'Stundenzettel KW{WEEKNUMBER} {NACHNAME}, {YEAR}',
    #                   f"""Hier ist der Stundenzettel für KW{WEEKNUMBER}.\n\nLiebe grüße, Rain's Bot."""
    #                   , files=output_path)

if __name__ == '__main__':

    print("Bearbeite und erstelle Daten ", end="", flush=True)
    progress_indicator = ProgressIndicator()
    progress_indicator.start()
    run()
    progress_indicator.stop()
    print("\rDaten erstellt und versendet.", end="")
